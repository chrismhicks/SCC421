import json
from nltk import word_tokenize
from openpyxl import Workbook
from openpyxl import load_workbook

rawJSONDataFile = "./Data/dataGathering/rawData.json"

saveFileName = "./Data/dataGathering/tokenizedData.xlsx"

# If the string contains a numerical value
def containsNumerical(string):
    for character in string:
        if (character.isnumeric()):
            return True
    return False


book = Workbook()
sheet = book.active
sheet["A1"], sheet["B1"], sheet["C1"], sheet["D1"] = "URL", "Sentence", "Token", "Name"
book.save(saveFileName)


bookToSave = load_workbook(saveFileName)
sheetToSave = bookToSave.active

# Parse through all the data collected from the webscraper and tokenise it at the word level
with open(rawJSONDataFile, "r") as jsonFile:
    for line in jsonFile:
        rawData = json.loads(line)
        sentenceNum = 0
        for sentence in rawData["Data"][0]["Summary"]:
            sentenceNum += 1
            for token in word_tokenize(sentence.lower()):
                # Only save as a token if its numerical or alphabetical
                if (token.isalpha() or containsNumerical(token)):
                    #print(token)
                    #print(rawData["Data"][0]["URL"])
                    #print(sentenceNum)
                    sheetToSave.append((rawData["Data"][0]["URL"], sentenceNum, token))
bookToSave.save(saveFileName)
