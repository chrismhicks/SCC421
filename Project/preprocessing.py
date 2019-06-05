from openpyxl import load_workbook
from sklearn import preprocessing
from sklearn.model_selection import KFold
import numpy as np
import loadData as ld
import random

def shortenData():
    workbook = load_workbook("./Data/dataGathering/tokenizedReducedData.xlsx")
    sheet = workbook.active

    data, labels = ld.loadData()
    numericalData = ld.loadNumericalTags()
    tags0 = ld.getTags(0)
    tags1 = ld.getTags(1)
    tags2 = ld.getTags(2)
    tags3 = ld.getTags(3)
    tags4 = ld.getTags(4)
    tags5 = ld.getTags(5)

    shortenedLabels, shortenedNumericalData, shortenedTags0, shortenedTags1, shortenedTags2, shortenedTags3, shortenedTags4, shortenedTags5 = [], [], [], [], [], [], [], []

    for i in range(0, len(labels)):
        if(labels[i] == "n" or labels[i] == "l" or labels[i] == "o"):
            shortenedLabels.append(labels[i])
            shortenedNumericalData.append(numericalData[i])
            shortenedTags0.append(tags0[i])
            shortenedTags1.append(tags1[i])
            shortenedTags2.append(tags2[i])
            shortenedTags3.append(tags3[i])
            shortenedTags4.append(tags4[i])
            shortenedTags5.append(tags5[i])


    for i in range(0, len(shortenedLabels)):
        sheet.cell(row=i+2, column=1).value = shortenedLabels[i]
        sheet.cell(row =i+2, column = 2).value = shortenedNumericalData[i]
        sheet.cell(row=i+2, column=3).value = shortenedTags0[i]
        sheet.cell(row=i+2, column=4).value = shortenedTags1[i]
        sheet.cell(row=i+2, column=5).value = shortenedTags2[i]
        sheet.cell(row=i+2, column=6).value = shortenedTags3[i]
        sheet.cell(row=i+2, column=7).value = shortenedTags4[i]
        sheet.cell(row=i+2, column=8).value = shortenedTags5[i]

    workbook.save("./Data/dataGathering/tokenizedReducedData.xlsx")


def getNormalisedData(combinedTags):
    #combinedTags = ld.loadCombinedTags()
    normalisedData = preprocessing.normalize(combinedTags)
    return normalisedData
    #print(combinedTags)
    #print(normalisedData)

def getStandardisedData(combinedTags):
    #combinedTags = ld.loadCombinedTags()
    standardisedData = preprocessing.scale(combinedTags)
    return standardisedData
    #print(combinedTags)
    #print(standardisedData)

#shortenData()

#getNormalisedData()
#getStandardisedData()

labels, posTags, tag0, tag1, tag2, tag3, tag4, tag5 = ld.loadShortenedData()

combinedReducedTags = []
for index in range(0, len(labels)):
    tempList = []
    tempList.append(tag0[index])
    tempList.append(tag1[index])
    tempList.append(tag2[index])
    tempList.append(tag3[index])
    tempList.append(tag4[index])
    tempList.append(tag5[index])
    combinedReducedTags.append(tempList)

#normalisedData = getNormalisedData(combinedReducedTags)

nIndex, lIndex, oIndex = [], [], []

for index in range(0, len(labels)):
    if(labels[index] == "n"):
        nIndex.append(index)
    elif(labels[index] == "l"):
        lIndex.append(index)
    elif(labels[index] == "o"):
        oIndex.append(index)

def oversample():
    for i in range(0, len(oIndex)-len(nIndex)):
        randN = random.randint(0, 794)
        nIndex.append(nIndex[randN])
    for i in range(0, len(oIndex) - len(lIndex)):
        randL = random.randint(0, 118)
        lIndex.append(lIndex[randL])

def crossValidation(numberOfFolds, data):
    trainSets, testSets = [], []

    # True to shuffle, 1 for RNG
    folds = KFold(numberOfFolds, True, 2)
    for train, test in folds.split(data):
        trainSets.append(train.tolist())
        testSets.append(test.tolist())
    return trainSets, testSets

oversample()

def prepareData(dq, folds):
    random.seed(a=15)
    # Choose how much data to use
    # Total per class is 1364 (total is 4092)
    # Number of each class
    dataQuantity = dq

    nData, lData, oData = [], [], []

    for i in range(0, dataQuantity):
        nData.append(nIndex[random.randint(0, len(nIndex)-1)])
        lData.append(lIndex[random.randint(0, len(lIndex)-1)])
        oData.append(oIndex[random.randint(0, len(oIndex)-1)])


    trainingSetsN, testingSetsN = crossValidation(folds, nData)
    trainingSetsL, testingSetsL = crossValidation(folds, lData)
    trainingSetsO, testingSetsO = crossValidation(folds, oData)

    trainingUsed, testingUsed = [], []
    for i in range(0, folds):
        tempTestN, tempTestL, tempTestO = [], [], []
        tempTrainN, tempTrainL, tempTrainO = [], [], []
        for j in range(0, len(trainingSetsN[i])):
            tempTestN.append(nData[trainingSetsN[i][j]])
            tempTestL.append(lData[trainingSetsL[i][j]])
            tempTestO.append(oData[trainingSetsO[i][j]])
        trainingUsed.append(tempTestN + tempTestL + tempTestO)

        for m in range(0, len(testingSetsN[i])):
            tempTrainN.append(nData[testingSetsN[i][m]])
            tempTrainL.append(lData[testingSetsL[i][m]])
            tempTrainO.append(oData[testingSetsO[i][m]])
        testingUsed.append(tempTrainN + tempTrainL + tempTrainO)

    return trainingUsed, testingUsed, labels, combinedReducedTags

def convertLabels(labels):
    newLabels = []
    for label in labels:
        if(label == "n"):
            newLabels.append(0)
        elif(label == "l"):
            newLabels.append(1)
        elif(label == "o"):
            newLabels.append(2)
        else:
            print("Incorrect Label: ", label)
    return newLabels